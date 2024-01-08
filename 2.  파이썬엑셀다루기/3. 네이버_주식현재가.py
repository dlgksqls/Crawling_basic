import requests
from bs4 import BeautifulSoup
import openpyxl

# 1) 엑셀 만들기

fpath = r'C:\Users\LeeHanBin\Desktop\crawling\Crawling_basic\2.  파이썬엑셀다루기\주식_현재가.xlsx'
# 2) 엑셀 워크시트 만들기
# wb = openpyxl.Workbook()
wb = openpyxl.load_workbook(fpath)

ws = wb.active # 현재 활성화 된 워크시트 참조

# 3) 데이터 추가하기

row = 2

codes = [
  '005930', '000660', '035720'
]

for code in codes:
  url = f"https://finance.naver.com/item/sise.naver?code={code}"
  response = requests.get(url)
  html = response.text
  soup = BeautifulSoup(html, 'html.parser')
  price = soup.select_one("#_nowVal").text
  price = price.replace(',','')
  ws[f'B{row}'] = int(price)
  row += 1
  
ws['A1'] = '종목'
ws['B1'] = '현재가'
ws['C1'] = '평균매입가'
ws['D1'] = '잔고수량'
ws['E1'] = '평가금액'
ws['F1'] = '평가손익'
ws['G1'] = '수익률'

ws['A2'] = '삼성전자'
ws['A3'] = 'SK하이닉스'
ws['A4'] = '카카오'

# 4) 엑셀 저장하기
wb.save(fpath)